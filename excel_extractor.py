# -*- coding: utf-8 -*-
"""
Created on Tue Oct 11 11:40:32 2022

@author: Benjamin Ojo
"""
# Import Packages
import openpyxl
import os

# File name.
cont_folder = 'C:/Users/User1/Documents/Data_Project/excel_data-wrangler/excel_files/container_file'

# Print list of file in a folder. 
files = os.listdir(cont_folder)

# Container code 
codes = []

for file in files:
    extr = file.split(' ')[1].split('.')[0]
    codes.append(extr)

# print(code)

# Creating extract file. 
saving_path = 'C:/Users/User1/Documents/Data_Project/excel_data-wrangler/extract.xlsx'


# Creating excel sheet. 
wb = openpyxl.workbook.Workbook()

for i in range(len(codes)):
    i = wb.create_sheet(codes[i])
    

# loading copy file
print('Loading coping sheet.', '\n')
for code in codes: 
    copy_path = f'C:/Users/User1/Documents/Data_Project/excel_data-wrangler/excel_files/container_file/CONTAINER {code}.xlsx'

    copy_wb = openpyxl.load_workbook(copy_path)
    copy_ws = copy_wb[code]

    past_ws = wb[code]

    # Get maximum row and columns for colume
    print(f'Getting Maximum row & column from data. Container-{code}', '\n')
    rm = copy_ws.max_row
    cm = copy_ws.max_column

    # coping sheet to extract folder. 
    print(f'Coping data from copy sheet. Container-{code}', '\n')
    for i in range(1, rm + 1): 
        for j in range(1, cm + 1):
            c = copy_ws.cell(row = i , column = j)
            
            past_ws.cell(row = i, column = j).value = c.value
    print(f'Coping sheet complete. Container-{code}', '\n\n')


wb.save(saving_path)
print('Saving Pasting sheet', '\n')

