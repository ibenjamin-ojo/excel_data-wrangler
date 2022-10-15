# -*- coding: utf-8 -*-
"""
Created on Tue Oct 11 11:40:32 2022

@author: Benjamin Ojo
"""
# Import Packages
import openpyxl
import os

# File name.
cont_folder = 'C:/Users/Services/OneDrive - IL Bagno Bathrooms + Design/Documents/Python Scripts/container_file'

# Print list of file in a folder. 
files = os.listdir(cont_folder)

# Container code 
code = []

for file in files:
    extr = file.split(' ')[1].split('.')[0]
    code.append(extr)

# print(code)

# Creating extract file. 
saving_path = 'C:/Users/Services/OneDrive - IL Bagno Bathrooms + Design/Documents/Python Scripts/extract.xlsx'


# Creating excel sheet. 
wb = openpyxl.workbook.Workbook()

for i in range(len(code)):
    i = wb.create_sheet(code[i])
    

# loading copy file
print('Loading coping sheet.', '\n')
copy_path = 'C:/Users/Services/OneDrive - IL Bagno Bathrooms + Design/Documents/Python Scripts/container_file/CONTAINER 196.xlsx'

copy_wb = openpyxl.load_workbook(copy_path)
copy_ws = copy_wb[code[0]]

past_ws = wb[code[0]]

# Get maximum row and columns for colume
print('Getting Maximum row & column from data.', '\n')
rm = copy_ws.max_row
cm = copy_ws.max_column

# coping sheet to extract folder. 
print('Coping data from copy sheet', '\n')
for i in range(1, rm + 1): 
    for j in range(1, cm + 1):
        c = copy_ws.cell(row = i , column = j)
        
        past_ws.cell(row = i, column = j).value = c.value
print('Coping sheet complete', '\n')

wb.save(saving_path)
print('Saving Pasting sheet', '\n')

